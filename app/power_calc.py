#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
アークナイツ キャラ火力計算ツール
各キャラのスキル火力（DPS・総ダメージ）を計算するCLIツール。
"""

import csv
import os
import sys
import re
import io
import datetime
import openpyxl

# UTF-8出力対応
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")
    sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8")

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(_SCRIPT_DIR, "..", "src"))
CSV_FILE = os.path.join(DATA_DIR, "arknights_star6.csv")
XLSX_DIR = os.path.join(DATA_DIR, "xlsx")
LOG_FILE = os.path.join(DATA_DIR, "calc_log.txt")

SKILL_RANK_ORDER = ["1", "2", "3", "4", "5", "6", "7", "特化I", "特化II", "特化III"]

RANK_DISPLAY = {
    "1": "ランク1", "2": "ランク2", "3": "ランク3",
    "4": "ランク4", "5": "ランク5", "6": "ランク6", "7": "ランク7",
    "特化I": "特化Ⅰ", "特化II": "特化Ⅱ", "特化III": "特化Ⅲ",
}


# ─── データ読み込み ──────────────────────────────────────────────

def parse_atk_speed(s: str) -> float:
    """'1.25s(やや遅い)' や '0.78s(とても速い)' などから秒数を抽出"""
    m = re.search(r"(\d+(?:\.\d+)?)s", s)
    return float(m.group(1)) if m else 1.0


def load_characters() -> list[dict]:
    """CSVからキャラクター情報を読み込む"""
    characters = []
    with open(CSV_FILE, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)  # ヘッダー行スキップ
        for row in reader:
            if len(row) < 13:
                continue
            name = row[1].strip()
            # テンプレート行・空行をスキップ
            if not name or name == "名前":
                continue
            try:
                atk = int(row[5])
                if atk == 0:
                    continue
            except ValueError:
                continue

            characters.append({
                "image":       row[0].strip(),
                "name":        name,
                "class":       row[2].strip(),
                "subclass":    row[3].strip(),
                "hp":          int(row[4]),
                "atk":         atk,
                "def":         int(row[6]),
                "res":         int(row[7]),
                "redeploy":    row[8].strip(),
                "cost":        int(row[9]),
                "block":       int(row[10]),
                "atk_speed":   parse_atk_speed(row[11]),
                "atk_speed_str": row[11].strip(),
                "source":      row[12].strip(),
                "tags":        row[13].strip() if len(row) > 13 else "",
            })
    return characters


# ─── スキルデータ解析 ────────────────────────────────────────────

def _closest_field(value, prev_init, prev_cost, prev_dur) -> str:
    """
    数値 value がどのフィールド(init/cost/dur)を表すか、
    前の状態との距離と変化方向で判定する。
    - init: 常に増加
    - cost: 常に減少
    - dur : 常に増加
    """
    dists = {}
    if not isinstance(value, (int, float)):
        return "cost"  # フォールバック

    if prev_init is not None and isinstance(prev_init, (int, float)) and value > prev_init:
        dists["init"] = abs(value - prev_init)

    if prev_cost is not None and isinstance(prev_cost, (int, float)) and value < prev_cost:
        dists["cost"] = abs(value - prev_cost)

    if prev_dur is not None and isinstance(prev_dur, (int, float)) and value > prev_dur:
        dists["dur"] = abs(value - prev_dur)

    if not dists:
        return "cost"  # フォールバック

    return min(dists, key=dists.get)


def _update_state(prev_init, prev_cost, prev_dur, numerics: list):
    """
    スキルシートの数値リスト(長さ0〜3)から state を更新する。
    numerics は (初期SP?, 必要SP?, 持続?) の順に並んでいるが、
    変化しなかったフィールドは省略(左詰め)されている。
    """
    init_sp, cost_sp, duration = prev_init, prev_cost, prev_dur

    if len(numerics) == 3:
        init_sp, cost_sp, duration = numerics[0], numerics[1], numerics[2]

    elif len(numerics) == 2:
        v1, v2 = numerics
        # v1 がどのフィールドか判定
        field1 = _closest_field(v1, prev_init, prev_cost, prev_dur)
        # v1 を除いた残りフィールドで v2 を判定
        rem_init = prev_init if field1 != "init" else None
        rem_cost = prev_cost if field1 != "cost" else None
        rem_dur  = prev_dur  if field1 != "dur"  else None
        field2 = _closest_field(v2, rem_init, rem_cost, rem_dur)

        if field1 == "init":  init_sp  = v1
        if field1 == "cost":  cost_sp  = v1
        if field1 == "dur":   duration = v1
        if field2 == "init":  init_sp  = v2
        if field2 == "cost":  cost_sp  = v2
        if field2 == "dur":   duration = v2

    elif len(numerics) == 1:
        v = numerics[0]
        field = _closest_field(v, prev_init, prev_cost, prev_dur)
        if field == "init":  init_sp  = v
        elif field == "cost": cost_sp = v
        elif field == "dur":  duration = v
        else:                 cost_sp  = v  # フォールバック

    return init_sp, cost_sp, duration


def parse_damage_multiplier(effect: str):
    """
    効果テキストから攻撃倍率を抽出する。
    戻り値: float (絶対倍率, 例 1.0=等倍, 2.0=200%)、解析失敗時は None

    対応パターン:
      - 攻撃力がX%まで/に上昇 → X/100
      - 攻撃力+X%  → 1 + X/100
      - 攻撃力、...(複数ステータス)+X%  → 1 + X/100
      - 攻撃力×X  → X
    """
    if not effect:
        return None

    # パターン1: 攻撃力がX%まで上昇 / 攻撃力がX%に上昇
    m = re.search(r"攻撃力[がが](\d+(?:\.\d+)?)%[まにに]で?上昇", effect)
    if m:
        return float(m.group(1)) / 100.0

    # パターン2: 攻撃力[任意文字]+X%
    # 「攻撃力+X%」や「攻撃力、防御力、最大HP+X%」のどちらにも対応
    m = re.search(r"攻撃力[^+\n]*?\+(\d+(?:\.\d+)?)%", effect)
    if m:
        return 1.0 + float(m.group(1)) / 100.0

    # パターン3: 攻撃力×X / 攻撃力xX
    m = re.search(r"攻撃力[××x](\d+(?:\.\d+)?)", effect)
    if m:
        return float(m.group(1))

    return None


def parse_skill_sheet(ws) -> dict:
    """
    スキル詳細シート(スキルN XXX1)を解析してランクごとのデータを返す。
    返値: {rank_str: {init_sp, cost_sp, duration, effect, multiplier}}
    """
    # シートから行データを取得
    row_data = []
    for row in ws.iter_rows():
        cells = {c.column_letter: c.value for c in row if c.value is not None}
        if cells:
            row_data.append(cells)

    if len(row_data) < 2:
        return {}

    # 初期状態
    init_sp = None
    cost_sp = None
    duration = None
    ranks = {}

    for cells in row_data[1:]:  # ヘッダー行スキップ
        rank = cells.get("A")
        if rank not in SKILL_RANK_ORDER:
            continue

        # B〜E の値を順に取得
        values = [cells[col] for col in ["B", "C", "D", "E"] if col in cells]

        # 効果テキスト(最後の文字列)と数値を分離
        effect_text = None
        numerics = []

        for v in values:
            s = str(v).strip()
            if s == "-":
                # '-' は持続なし(パッシブ/瞬時)を意味する
                numerics.append(None)
            else:
                try:
                    n = float(s) if "." in s else int(s)
                    numerics.append(n)
                except ValueError:
                    effect_text = s

        # '-' が持続として入ってきた場合のみ duration に反映
        # (数値が3つある場合は init,cost,dur の順)
        if len(numerics) == 3 and numerics[2] is None:
            # init,cost は通常通り、dur=None (持続なし)
            init_sp = numerics[0] if numerics[0] is not None else init_sp
            cost_sp = numerics[1] if numerics[1] is not None else cost_sp
            duration = None
        elif None in numerics:
            # '-'が混在する場合: duration=None として残りを処理
            valid_nums = [n for n in numerics if n is not None]
            init_sp, cost_sp, duration = _update_state(init_sp, cost_sp, duration, valid_nums)
            if numerics[-1] is None or (len(numerics) >= 1 and numerics[0] is None and len(valid_nums) == 0):
                duration = None
        else:
            init_sp, cost_sp, duration = _update_state(init_sp, cost_sp, duration, numerics)

        ranks[str(rank)] = {
            "init_sp":    init_sp,
            "cost_sp":    cost_sp,
            "duration":   duration,
            "effect":     effect_text,
            "multiplier": parse_damage_multiplier(effect_text),
        }

    return ranks


def load_skills(char_name: str) -> list[dict] | None:
    """キャラ名に対応する xlsx からスキル情報を読み込む"""
    xlsx_path = os.path.join(XLSX_DIR, f"{char_name}.xlsx")
    if not os.path.exists(xlsx_path):
        return None

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"  [警告] {xlsx_path} の読み込みに失敗: {e}")
        return None

    skills = []
    for sheet_name in wb.sheetnames:
        # "スキルN 名前1" (末尾が数字1) → 詳細データシート
        m = re.match(r"スキル(\d+) (.+?)1$", sheet_name)
        if not m:
            continue
        skill_num  = int(m.group(1))
        skill_name = m.group(2)
        ws = wb[sheet_name]
        ranks = parse_skill_sheet(ws)
        skills.append({
            "num":   skill_num,
            "name":  skill_name,
            "ranks": ranks,
        })

    skills.sort(key=lambda x: x["num"])
    return skills


# ─── ダメージ計算 ────────────────────────────────────────────────

def calc_damage(atk: int, multiplier: float, enemy_def: int,
                enemy_res: int, is_arts: bool) -> tuple[float, float]:
    """
    ダメージを計算する。
    返値: (スキル発動時のダメージ(軽減前), 実ダメージ(軽減後))

    アークナイツのダメージ計算式:
      物理: max(ATK × mult - DEF, ATK × mult × 0.05)
      術 :  ATK × mult × max(1 - RES/100, 0.05)
    """
    raw = atk * multiplier
    if is_arts:
        reduction = min(enemy_res / 100.0, 0.95)
        actual = raw * (1.0 - reduction)
    else:
        actual = max(raw - enemy_def, raw * 0.05)
    return raw, actual


def calc_total_damage(actual_per_hit: float, duration,
                      atk_speed: float, targets: int = 1) -> float | None:
    """スキル継続中の総ダメージを計算する"""
    if duration is None or not isinstance(duration, (int, float)) or duration <= 0:
        return None
    hits = int(float(duration) / atk_speed)
    return actual_per_hit * hits * targets


# ─── ユーティリティ ──────────────────────────────────────────────

def input_int(prompt: str, default: int) -> int:
    while True:
        try:
            s = input(f"{prompt} [デフォルト: {default}]: ").strip()
            return int(s) if s else default
        except ValueError:
            print("  整数を入力してください。")


def fmt_sp(val) -> str:
    """SP/持続値を表示用文字列に変換（0はそのまま表示）"""
    if val is None:
        return "-"
    return str(val)


def select_from_list(items: list, prompt: str, display_fn=None) -> int:
    """リストから番号選択。0-indexedで返す"""
    for i, item in enumerate(items, 1):
        label = display_fn(item) if display_fn else str(item)
        print(f"  {i:3}. {label}")
    while True:
        s = input(f"{prompt}: ").strip()
        if s.isdigit():
            idx = int(s) - 1
            if 0 <= idx < len(items):
                return idx
        print(f"  1〜{len(items)} の番号を入力してください。")


# ─── メイン ─────────────────────────────────────────────────────

def calc_session(characters: list[dict]):
    """1回分の計算セッション"""

    # ─── キャラクター選択 ───
    print("\n【キャラクター選択】")

    def char_label(c):
        return f"{c['name']:16s}  {c['class']:4s} / {c['subclass']:10s}  ATK:{c['atk']:4d}  速度:{c['atk_speed_str']}"

    # キャラ名またはインデックスで検索
    while True:
        query = input("\nキャラ名または番号を入力 (一覧は 'list'): ").strip()
        if query.lower() == "list":
            for i, c in enumerate(characters, 1):
                print(f"  {i:3}. {char_label(c)}")
            continue

        char = None
        if query.isdigit():
            idx = int(query) - 1
            if 0 <= idx < len(characters):
                char = characters[idx]
        else:
            # 部分一致検索
            matches = [c for c in characters if query in c["name"]]
            if len(matches) == 1:
                char = matches[0]
            elif len(matches) > 1:
                print(f"  {len(matches)} 件ヒットしました:")
                idx = select_from_list(matches, "番号を選択", lambda c: c["name"])
                char = matches[idx]
        if char:
            break
        print("  キャラが見つかりませんでした。もう一度入力してください。")

    print(f"\n  選択: {char['name']}")
    print(f"  職業: {char['class']} / {char['subclass']}")
    print(f"  攻撃力: {char['atk']}  攻撃速度: {char['atk_speed']}s  HP: {char['hp']}")

    # ─── スキルデータ読み込み ───
    print("\nスキルデータを読み込み中...")
    skills = load_skills(char["name"])

    if not skills:
        print("  スキルデータが見つかりません。")
        return

    print(f"  {len(skills)} スキル読み込み完了\n")

    # ─── スキル選択 ───
    print("【スキル選択】")
    for s in skills:
        print(f"  {s['num']}. {s['name']}")

    while True:
        s_input = input("スキル番号を入力 (1〜3): ").strip()
        if s_input.isdigit():
            num = int(s_input)
            sel_skills = [s for s in skills if s["num"] == num]
            if sel_skills:
                skill = sel_skills[0]
                break
        print("  有効なスキル番号を入力してください。")

    print(f"\n  選択: スキル{skill['num']} {skill['name']}")

    # ─── ランク選択 ───
    print("\n【ランク選択】")
    available = [r for r in SKILL_RANK_ORDER if r in skill["ranks"]]

    if not available:
        print("  ランクデータがありません。")
        return

    for r in available:
        rd = skill["ranks"][r]
        mult = rd.get("multiplier")
        mult_str = f"  倍率:{mult:.0%}" if mult else ""
        dur = rd.get("duration")
        dur_str = f"  持続:{dur}s" if isinstance(dur, (int, float)) else "  持続:なし"
        print(f"  {RANK_DISPLAY[r]:6s}  SP初期:{fmt_sp(rd['init_sp']):>4}  SP必要:{fmt_sp(rd['cost_sp']):>4}"
              f"{dur_str}{mult_str}")

    while True:
        r_input = input("\nランクを入力 (例: 7 / 特化I / 特化II / 特化III): ").strip()
        # 入力の正規化
        r_norm = r_input.replace("Ⅰ", "I").replace("Ⅱ", "II").replace("Ⅲ", "III")
        if r_norm in skill["ranks"]:
            selected_rank = r_norm
            break
        # 短縮入力サポート: "3" → "特化III", "m3" → "特化III"
        alias = {
            "m1": "特化I", "m2": "特化II", "m3": "特化III",
            "1": "1", "2": "2", "3": "3", "4": "4",
            "5": "5", "6": "6", "7": "7",
        }
        if r_input in alias and alias[r_input] in skill["ranks"]:
            selected_rank = alias[r_input]
            break
        print("  有効なランクを入力してください。")

    rank_data = skill["ranks"][selected_rank]

    print(f"\n  選択ランク: {RANK_DISPLAY[selected_rank]}")
    if rank_data["effect"]:
        print(f"  効果: {rank_data['effect'][:80]}...")

    # ─── ダメージ種別 ───
    print("\n【ダメージ種別】")

    # 効果テキストから自動判定
    arts_auto = (
        char["class"] == "術師"
        or (rank_data["effect"] and "術ダメージ" in rank_data["effect"])
    )

    print(f"  1. 物理ダメージ{'  ← 推定' if not arts_auto else ''}")
    print(f"  2. 術ダメージ{'  ← 推定' if arts_auto else ''}")

    while True:
        d_input = input(f"ダメージ種別を選択 [デフォルト: {'2' if arts_auto else '1'}]: ").strip()
        if d_input == "":
            is_arts = arts_auto
            break
        if d_input in ("1", "2"):
            is_arts = (d_input == "2")
            break
        print("  1 または 2 を入力してください。")

    # ─── 敵ステータス入力 ───
    print("\n【敵ステータス】")
    if is_arts:
        enemy_res = input_int("敵の術耐性 (0〜100)", 0)
        enemy_def = 0
    else:
        enemy_def = input_int("敵の防御力", 300)
        enemy_res = 0

    # ─── 攻撃対象数 ───
    targets = input_int("攻撃対象数 (スキル中の同時攻撃数)", 1)

    # ─── 倍率確認 / 手動入力 ───
    multiplier = rank_data.get("multiplier")

    if multiplier is None:
        print("\n  [注意] 攻撃倍率を効果テキストから自動解析できませんでした。")
        if rank_data["effect"]:
            print(f"  効果テキスト: {rank_data['effect']}")
        while True:
            try:
                s = input("  攻撃倍率を手動で入力 (例: 3.30 = 330%): ").strip()
                multiplier = float(s)
                break
            except ValueError:
                print("  数値を入力してください (例: 3.30)")

    # ─── 持続時間の確認 / 手動入力 ───
    duration = rank_data.get("duration")
    if not isinstance(duration, (int, float)) or duration <= 0:
        print(f"\n  [情報] 持続時間がデータにありません (持続:'-' またはデータ不足)")
        dur_input = input("  持続時間を手動で入力 (例: 40  / スキップはEnter): ").strip()
        if dur_input:
            try:
                duration = float(dur_input)
            except ValueError:
                duration = None

    # ─── 計算 ───
    raw_dmg, actual_dmg = calc_damage(
        char["atk"], multiplier, enemy_def, enemy_res, is_arts
    )

    total_dmg = calc_total_damage(actual_dmg, duration, char["atk_speed"], targets)

    hits = None
    if isinstance(duration, (int, float)) and duration > 0:
        hits = int(duration / char["atk_speed"])

    # ─── 結果表示 & ログ記録 ───
    lines = []
    lines.append("")
    lines.append("=" * 60)
    lines.append("  ===  火力計算結果  ===")
    lines.append("=" * 60)
    lines.append(f"  キャラクター : {char['name']}")
    lines.append(f"  スキル       : スキル{skill['num']} {skill['name']} [{RANK_DISPLAY[selected_rank]}]")
    lines.append(f"  攻撃力       : {char['atk']}")
    lines.append(f"  攻撃速度     : {char['atk_speed']}s / hit")
    lines.append(f"  ダメージ種別 : {'術ダメージ' if is_arts else '物理ダメージ'}")
    if is_arts:
        lines.append(f"  敵の術耐性   : {enemy_res}%")
    else:
        lines.append(f"  敵の防御力   : {enemy_def}")
    lines.append(f"  攻撃倍率     : {multiplier:.0%}  ({multiplier:.2f}x)")
    lines.append("-" * 60)

    lines.append("")
    lines.append("  【スキル発動時のダメージ (軽減前)】")
    lines.append(f"    1発あたり : {raw_dmg:,.0f}")
    if targets > 1:
        lines.append(f"    {targets}体同時   : {raw_dmg * targets:,.0f}")

    lines.append("")
    lines.append("  【実ダメージ (軽減後)】")
    if is_arts:
        eff_red = min(enemy_res, 95)
        lines.append(f"    術耐性軽減後 ({eff_red}%軽減): {actual_dmg:,.0f}")
    else:
        lines.append(f"    防御力軽減後 ({enemy_def}防御): {actual_dmg:,.0f}")
    if targets > 1:
        lines.append(f"    {targets}体合計        : {actual_dmg * targets:,.0f}")

    lines.append("")
    lines.append("  【スキル継続中の総ダメージ】")
    if isinstance(duration, (int, float)) and duration > 0:
        lines.append(f"    持続時間   : {duration}s")
        lines.append(f"    ヒット数   : {hits} 回 × {targets} 体")
        lines.append(f"    総ダメージ : {total_dmg:,.0f}")
        dps = actual_dmg / char["atk_speed"] * targets
        lines.append(f"    スキル中DPS: {dps:,.1f} / s")
    else:
        lines.append("    持続時間なし（瞬時発動・パッシブ型）")
        lines.append(f"    ※ 通常攻撃DPS = {char['atk'] / char['atk_speed']:,.1f} / s")

    # SP効率参考
    init = rank_data.get("init_sp")
    cost = rank_data.get("cost_sp")
    lines.append("")
    lines.append("  【SP情報 (参考)】")
    lines.append(f"    初期SP: {fmt_sp(init)}  /  必要SP: {fmt_sp(cost)}")

    lines.append("")
    lines.append("=" * 60)

    # 画面出力
    for line in lines:
        print(line)

    # ファイル出力（追記）
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"\n[{timestamp}]\n")
            f.write("\n".join(lines))
            f.write("\n")
        print(f"  [ログ保存] {LOG_FILE}")
    except OSError as e:
        print(f"  [警告] ログ保存に失敗しました: {e}")


def main():
    """エントリーポイント: データ読み込み → 計算ループ"""
    print("=" * 60)
    print("  アークナイツ キャラ火力計算ツール")
    print("=" * 60)

    print("\nキャラクターデータを読み込み中...")
    characters = load_characters()
    print(f"  {len(characters)} キャラクター読み込み完了")

    while True:
        try:
            calc_session(characters)
        except (KeyboardInterrupt, EOFError):
            print("\n\n終了します。")
            break
        try:
            again = input("\n別の計算をしますか？ (y/N): ").strip().lower()
        except (KeyboardInterrupt, EOFError):
            print("\n終了します。")
            break
        if again != "y":
            print("終了します。")
            break


if __name__ == "__main__":
    main()
